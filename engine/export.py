"""
Report exporters — PDF (reportlab) and XLSX (openpyxl).

Both exporters are pure functions of the report dict produced by
engine.checker.run_ruleset. No timestamps, no randomness: identical
report dicts always produce identical documents (determinism is the
product's core claim). The report SHA-256 embedded in both formats
lets a reviewer verify the exported document matches a given report.
"""

from __future__ import annotations

import datetime
import hashlib
import json

# ---------------------------------------------------------------------------
# Shared

STATUS_ORDER = ["pass", "fail", "info_not_available", "uncertain"]

# One palette for both formats (hex without '#'):
# pass green, fail red, uncertain amber, info_not_available grey.
STATUS_COLORS = {
    "pass": "C6EFCE",
    "fail": "FFC7CE",
    "uncertain": "FFE699",
    "info_not_available": "D9D9D9",
}
STATUS_TEXT_COLORS = {
    "pass": "1E7B34",
    "fail": "B00020",
    "uncertain": "9C6500",
    "info_not_available": "595959",
}


def report_sha256(report: dict) -> str:
    """SHA-256 hexdigest of the canonical (sort_keys) JSON serialization."""
    return hashlib.sha256(
        json.dumps(report, sort_keys=True).encode()
    ).hexdigest()


def _fact_str(f: dict) -> str:
    conf = f.get("confidence")
    conf_s = "n/a" if conf is None else conf
    source = f.get("source") or "n/a"
    return f"{f['fact']}={f['value']} @conf {conf_s} [{source}]"


def _facts_used_cell(facts_used: list[dict]) -> str:
    return "; ".join(_fact_str(f) for f in facts_used)


# ---------------------------------------------------------------------------
# XLSX

def to_xlsx(report: dict, path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # openpyxl stamps docProps with datetime.now() by default; pin them to a
    # fixed epoch so identical reports produce byte-identical files
    # (determinism — this is a constant, not a wall-clock read).
    epoch = datetime.datetime(2020, 1, 1)
    wb.properties.created = epoch
    wb.properties.modified = epoch

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    bold = Font(bold=True)

    project = report.get("project", {})
    rows: list[tuple[str, str]] = [("Project", project.get("name", "?"))]
    if "municipality" in project:
        rows.append(("Municipality", project["municipality"]))
    rows.append(("Ruleset", report.get("ruleset_id", "?")))
    rows.append(("Code edition", report.get("code_edition", "?")))

    summary = report.get("summary", {})
    for status in STATUS_ORDER:
        if status in summary:
            rows.append((f"Checks: {status}", summary[status]))
    for status in summary:  # any statuses beyond the known four
        if status not in STATUS_ORDER:
            rows.append((f"Checks: {status}", summary[status]))

    engine = report.get("engine", {})
    if engine.get("note"):
        rows.append(("Engine note", engine["note"]))
    rows.append(("Report SHA-256", report_sha256(report)))

    for label, value in rows:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = bold
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    # --- Checks sheet ---
    ws = wb.create_sheet("Checks")
    header = ["Rule ID", "Provision", "Title", "Entity", "Status",
              "Detail", "Facts Used", "Comparisons"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = bold

    for r in report.get("results", []):
        status = r["status"]
        ws.append([
            r["rule_id"],
            r["provision"],
            r["title"],
            f"{r['entity_name']} ({r['entity_id']})",
            status,
            r["detail"],
            _facts_used_cell(r.get("facts_used", [])),
            "; ".join(r.get("comparisons", [])),
        ])
        fill_hex = STATUS_COLORS.get(status)
        if fill_hex:
            ws.cell(row=ws.max_row, column=5).fill = PatternFill(
                start_color=fill_hex, end_color=fill_hex, fill_type="solid"
            )

    ws.freeze_panes = "A2"
    widths = [26, 40, 34, 30, 18, 60, 70, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    _normalize_zip_timestamps(path)


def _normalize_zip_timestamps(path: str) -> None:
    """Rewrite the xlsx (a zip) with fixed timestamps.

    Two wall-clock leaks make otherwise-identical exports differ: the zip
    container stamps every member with mtime, and openpyxl overwrites
    docProps/core.xml <dcterms:modified> with now() at save time. Pin both
    to a constant so identical reports produce byte-identical files.
    """
    import re
    import zipfile

    fixed = (2020, 1, 1, 0, 0, 0)
    fixed_w3cdtf = "2020-01-01T00:00:00Z"
    with zipfile.ZipFile(path) as zin:
        members = [(item, zin.read(item.filename)) for item in zin.infolist()]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item, data in members:
            if item.filename == "docProps/core.xml":
                data = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>" + fixed_w3cdtf.encode() + rb"\g<2>",
                    data,
                )
            info = zipfile.ZipInfo(item.filename, date_time=fixed)
            info.external_attr = item.external_attr
            info.compress_type = zipfile.ZIP_DEFLATED
            zout.writestr(info, data)


# ---------------------------------------------------------------------------
# PDF

def to_pdf(report: dict, path: str) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        ListFlowable, ListItem, Paragraph, SimpleDocTemplate, Spacer, Table,
        TableStyle,
    )

    styles = getSampleStyleSheet()
    body = styles["BodyText"]
    small = ParagraphStyle("small", parent=body, fontSize=8, leading=10)
    italic = styles["Italic"]

    def esc(s: object) -> str:
        return (str(s).replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;"))

    project_name = report.get("project", {}).get("name", "?")
    sha = report_sha256(report)

    story = [
        Paragraph(f"NBC Compliance Check — {esc(project_name)}",
                  styles["Title"]),
        Paragraph(
            f"Ruleset: {esc(report.get('ruleset_id', '?'))} — "
            f"{esc(report.get('code_edition', '?'))}",
            styles["Heading2"]),
        Spacer(1, 4 * mm),
    ]

    # Summary table (status counts)
    summary = report.get("summary", {})
    sum_rows = [["Status", "Count"]]
    for status in STATUS_ORDER:
        if status in summary:
            sum_rows.append([status, str(summary[status])])
    for status in summary:
        if status not in STATUS_ORDER:
            sum_rows.append([status, str(summary[status])])
    sum_table = Table(sum_rows, colWidths=[60 * mm, 25 * mm], hAlign="LEFT")
    sum_style = [
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEEEEE")),
    ]
    for i, row in enumerate(sum_rows[1:], start=1):
        hexcol = STATUS_COLORS.get(row[0])
        if hexcol:
            sum_style.append(
                ("BACKGROUND", (0, i), (0, i), colors.HexColor(f"#{hexcol}")))
    sum_table.setStyle(TableStyle(sum_style))
    story.append(sum_table)
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph(
        "Deterministic engine — identical inputs always produce identical "
        f"results. Report SHA-256: {sha}", body))
    story.append(Spacer(1, 6 * mm))

    # Per-check sections
    for r in report.get("results", []):
        status = r["status"]
        color = STATUS_TEXT_COLORS.get(status, "000000")
        story.append(Paragraph(
            f'<font color="#{color}">[{esc(status.upper())}]</font> '
            f"{esc(r['rule_id'])} — {esc(r['title'])}",
            styles["Heading3"]))
        story.append(Paragraph(
            f"{esc(r['provision'])} — entity: {esc(r['entity_name'])} "
            f"({esc(r['entity_id'])})", italic))
        story.append(Paragraph(esc(r["detail"]), body))

        facts_used = r.get("facts_used", [])
        if facts_used:
            fact_rows = [[
                Paragraph("<b>Fact</b>", small),
                Paragraph("<b>Value</b>", small),
                Paragraph("<b>Confidence</b>", small),
                Paragraph("<b>Source</b>", small),
            ]]
            for f in facts_used:
                conf = f.get("confidence")
                fact_rows.append([
                    Paragraph(esc(f["fact"]), small),
                    Paragraph(esc(f.get("value")), small),
                    Paragraph("n/a" if conf is None else esc(conf), small),
                    Paragraph(esc(f.get("source") or "n/a"), small),
                ])
            ft = Table(fact_rows,
                       colWidths=[38 * mm, 28 * mm, 24 * mm, 80 * mm],
                       hAlign="LEFT")
            ft.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEEEEE")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(Spacer(1, 2 * mm))
            story.append(ft)

        comparisons = r.get("comparisons", [])
        if comparisons:
            story.append(Spacer(1, 2 * mm))
            story.append(ListFlowable(
                [ListItem(Paragraph(esc(c), small)) for c in comparisons],
                bulletType="bullet", leftIndent=8 * mm))
        story.append(Spacer(1, 4 * mm))

    # invariant=1 strips reportlab's build timestamp so identical reports
    # produce byte-identical PDFs.
    doc = SimpleDocTemplate(
        path, pagesize=letter, invariant=1,
        title=f"NBC Compliance Check — {project_name}",
        author="nbc-checker deterministic engine",
    )
    doc.build(story)


# ---------------------------------------------------------------------------
# CLI

def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(
        description="Export an engine report JSON to PDF and/or XLSX.")
    parser.add_argument("report", help="Path to report JSON (e.g. reports/last_report.json)")
    parser.add_argument("--pdf", help="Write a PDF report to this path")
    parser.add_argument("--xlsx", help="Write an XLSX report to this path")
    args = parser.parse_args()

    if not args.pdf and not args.xlsx:
        parser.error("nothing to do: provide --pdf and/or --xlsx")

    with open(args.report) as f:
        report = json.load(f)

    if args.xlsx:
        to_xlsx(report, args.xlsx)
        print(f"wrote {args.xlsx}")
    if args.pdf:
        to_pdf(report, args.pdf)
        print(f"wrote {args.pdf}")
    print(f"report sha256: {report_sha256(report)}")


if __name__ == "__main__":
    main()
