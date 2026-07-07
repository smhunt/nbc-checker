"""Exporter tests — XLSX/PDF outputs and report hashing."""
import copy

from engine.checker import run_ruleset
from engine.export import report_sha256, to_pdf, to_xlsx

RULESET = {
    "ruleset_id": "test-ruleset",
    "code_edition": "NBC 2020 (test)",
    "rules": [
        {
            "rule_id": "T-RISE", "provision": "TEST 1.1.1.",
            "title": "Riser height",
            "applies_to": {"entity_type": "stair", "where": {}},
            "requires": {"all": [
                {"fact": "riser_height_mm", "op": "<=", "value": 200},
            ]},
            "exceptions": [],
        },
        {
            "rule_id": "T-RUN", "provision": "TEST 2.2.2.",
            "title": "Tread run",
            "applies_to": {"entity_type": "stair", "where": {}},
            "requires": {"all": [
                {"fact": "tread_run_mm", "op": ">=", "value": 255},
            ]},
            "exceptions": [],
        },
        {
            "rule_id": "T-WIDTH", "provision": "TEST 3.3.3.",
            "title": "Clear width",
            "applies_to": {"entity_type": "stair", "where": {}},
            "requires": {"all": [
                {"fact": "clear_width_mm", "op": ">=", "value": 860},
            ]},
            "exceptions": [],
        },
    ],
}

FACTS = {
    "project": {"name": "Export Test Dwelling", "municipality": "Testville"},
    "entities": [
        {
            "entity_type": "stair", "id": "s1", "name": "Stair 1",
            "attributes": {
                "riser_height_mm": 185,  # pass
                "tread_run_mm": 250,     # fail (< 255)
                "clear_width_mm": {      # uncertain (confidence < 0.9)
                    "value": 900, "confidence": 0.7,
                    "source": "plans.pdf p.3",
                },
            },
        },
    ],
}


def make_report():
    return run_ruleset(RULESET, FACTS)


def test_report_has_expected_statuses():
    report = make_report()
    statuses = {r["rule_id"]: r["status"] for r in report["results"]}
    assert statuses == {
        "T-RISE": "pass", "T-RUN": "fail", "T-WIDTH": "uncertain",
    }


def test_to_xlsx(tmp_path):
    from openpyxl import load_workbook

    report = make_report()
    path = str(tmp_path / "out.xlsx")
    to_xlsx(report, path)

    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"Summary", "Checks"}

    checks = wb["Checks"]
    rows = list(checks.iter_rows(values_only=True))
    assert len(rows) == len(report["results"]) + 1  # header + one per result
    assert rows[0][0] == "Rule ID"
    assert rows[1][0] == "T-RISE"
    assert rows[1][4] == "pass"
    fail_row = next(r for r in rows[1:] if r[0] == "T-RUN")
    assert fail_row[4] == "fail"
    # Facts Used cell format, including confidence/source rendering
    uncertain_row = next(r for r in rows[1:] if r[0] == "T-WIDTH")
    assert "clear_width_mm=900 @conf 0.7 [plans.pdf p.3]" in uncertain_row[6]

    # Summary counts match the report's summary
    summary = wb["Summary"]
    summary_map = {
        row[0]: row[1] for row in summary.iter_rows(values_only=True)
    }
    for status, count in report["summary"].items():
        assert summary_map[f"Checks: {status}"] == count
    assert summary_map["Project"] == "Export Test Dwelling"
    assert summary_map["Municipality"] == "Testville"
    assert summary_map["Report SHA-256"] == report_sha256(report)


def test_to_pdf(tmp_path):
    import os

    report = make_report()
    path = str(tmp_path / "out.pdf")
    to_pdf(report, path)

    assert os.path.exists(path)
    assert os.path.getsize(path) > 1500
    with open(path, "rb") as f:
        data = f.read()
    assert b"NBC Compliance Check" in data


def test_report_sha256_deterministic_and_sensitive():
    report = make_report()
    assert report_sha256(report) == report_sha256(make_report())

    changed = copy.deepcopy(report)
    changed["results"][0]["status"] = "fail"
    assert report_sha256(changed) != report_sha256(report)
